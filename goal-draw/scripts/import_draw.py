# Imports the draw from Nathaniel's spreadsheet into the app's static teams.json.
# Run again whenever the spreadsheet changes (e.g. when Status flips to FINAL):
#   python import_draw.py
# It is a one-off snapshot — the live site never reads the spreadsheet.

import json
import sys
from pathlib import Path

import openpyxl

XLSX = Path(r"C:\Users\natha\Documents\Misc\Prem Goal Draw Randomiser.xlsx")
OUT = Path(__file__).resolve().parent.parent / "teams.json"

# 3-letter codes for every club that can appear in the draw or as a stand-in,
# keyed by the name used on the spreadsheet's Badge tab.
ABBR = {
    "Arsenal": "ARS", "Manchester City": "MCI", "AFC Bournemouth": "BOU",
    "Manchester United": "MUN", "Liverpool": "LIV", "Aston Villa": "AVL",
    "Chelsea": "CHE", "Newcastle United": "NEW", "Brighton & Hove Albion": "BHA",
    "Brentford": "BRE", "Fulham": "FUL", "Leeds United": "LEE", "Everton": "EVE",
    "Nottingham Forest": "NFO", "Tottenham Hotspur": "TOT", "Sunderland": "SUN",
    "Crystal Palace": "CRY", "Hull City": "HUL", "Ipswich Town": "IPS",
    "Coventry City": "COV", "West Ham United": "WHU", "Burnley": "BUR",
    "Wolverhampton": "WOL",
}

# Kit chip designs, interpreted from the Badge tab's kit descriptions.
# style: plain | stripes (vertical, body/body2) | sleeves (contrast sleeves) |
#        pinstripe | shoulders | sleeve-trim
KITS = {
    "ARS": {"style": "sleeves", "body": "#EF0107", "sleeves": "#FFFFFF"},
    "MCI": {"style": "plain", "body": "#6CABDD"},
    "BOU": {"style": "stripes", "body": "#DA020E", "body2": "#000000", "sleeves": "#000000"},
    "MUN": {"style": "plain", "body": "#DA291C"},
    "LIV": {"style": "plain", "body": "#C8102E"},
    "AVL": {"style": "sleeves", "body": "#670E36", "sleeves": "#95BFE5"},
    "CHE": {"style": "plain", "body": "#034694"},
    "NEW": {"style": "stripes", "body": "#241F20", "body2": "#FFFFFF", "sleeves": "#241F20"},
    "BHA": {"style": "stripes", "body": "#0057B8", "body2": "#FFFFFF", "sleeves": "#0057B8"},
    "BRE": {"style": "stripes", "body": "#E30613", "body2": "#FFFFFF", "sleeves": "#FFFFFF"},
    "FUL": {"style": "sleeve-trim", "body": "#FFFFFF", "trim": "#000000", "dark_text": True},
    "LEE": {"style": "sleeve-trim", "body": "#FFFFFF", "trim": "#1D428A", "dark_text": True},
    "EVE": {"style": "plain", "body": "#003399"},
    "NFO": {"style": "pinstripe", "body": "#DD0000", "pin": "#FFFFFF"},
    "TOT": {"style": "shoulders", "body": "#FFFFFF", "shoulders": "#132257", "dark_text": True},
    "SUN": {"style": "stripes", "body": "#EB172B", "body2": "#FFFFFF", "sleeves": "#EB172B"},
    "CRY": {"style": "stripes", "body": "#C4122E", "body2": "#1B458F", "sleeves": "#1B458F"},
    "HUL": {"style": "stripes", "body": "#F5971D", "body2": "#000000", "sleeves": "#000000"},
    "IPS": {"style": "plain", "body": "#3A64A3"},
    "COV": {"style": "stripes", "body": "#63B1E5", "body2": "#FFFFFF", "sleeves": "#63B1E5"},
    "WHU": {"style": "plain", "body": "#7A263A"},
    "BUR": {"style": "plain", "body": "#6C1D45"},
    "WOL": {"style": "plain", "body": "#FDB913", "dark_text": True},
}

# Promoted clubs have no 2025-26 Premier League data, so the pre-season replay
# uses the relegated club that vacated their slot as a stand-in.
STANDINS = {"HUL": "WOL", "COV": "WHU", "IPS": "BUR"}


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    draw = wb["Final Draw"]
    status = None
    players = {}
    header_seen = False
    for row in draw.iter_rows(values_only=True):
        vals = [v for v in row]
        if any(v == "Status:" for v in vals):
            status = str(vals[vals.index("Status:") + 1]).strip().upper()
            continue
        if "Player Name" in vals:
            header_seen = True
            continue
        if header_seen and any(v for v in vals):
            cells = [str(v).strip() for v in vals if v not in (None, "")]
            name = cells[-1]
            teams = cells[:-1]
            missing = [t for t in teams if t not in ABBR]
            if missing:
                sys.exit(f"Unknown team(s) in draw for {name}: {missing}")
            players[name] = [ABBR[t] for t in teams]

    badge = wb["Badge"]
    teams = {}
    for row in badge.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name, abbr = str(row[0]).strip(), str(row[1]).strip().upper()
        if abbr not in KITS:
            sys.exit(f"No kit design for {abbr} ({name}) — add one to KITS.")
        display = "Wolverhampton Wanderers" if name == "Wolverhampton" else name
        teams[abbr] = {"name": display, "kit": KITS[abbr]}

    out = {
        "season": "2026-27",
        "kickoff": "2026-08-21T20:00:00+01:00",
        "status": status or "DRAFT",
        "players": players,
        "standins": STANDINS,
        "teams": teams,
    }
    OUT.write_text(json.dumps(out, indent=2), encoding="utf-8")
    print(f"Wrote {OUT} — status {out['status']}, "
          f"{len(players)} players, {len(teams)} team badges")


if __name__ == "__main__":
    main()
